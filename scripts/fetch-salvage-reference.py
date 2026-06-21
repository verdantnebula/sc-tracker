# ============================================================================
# scripts/fetch-salvage-reference.py — DEV-ONLY salvage reference generator
# ----------------------------------------------------------------------------
# Regenerates electron/data/salvage-reference.json from a community salvage
# worksheet (an .xlsx). Run this once per game patch (ship loadouts, component
# prices and material rates change per patch):
#
#     python scripts/fetch-salvage-reference.py "<path to worksheet.xlsx>"
#     # or set the default path below and run with no arg
#
# Requires openpyxl (`pip install openpyxl`). The worksheet itself is NEVER
# copied into the repo — only the sanitized JSON this script emits is bundled.
# The packaged app reads only that JSON; it makes no network calls and ships no
# token. This is purely game reference data (ship/component names + prices) —
# no personal data.
#
# Worksheet layout this parser targets (SC 4.8 community salvage worksheet):
#   - 'Salvage Tracking' : ships grouped under "Cost : N (M)" headers. Each block
#       is a horizontal table — a 'Ship' row naming ships across columns, then
#       attribute rows (Cargo SCU / CMAT / Power Plant / Shield / Quantum Drive /
#       Cooler / Radar / Weapon), each column = one ship. Weapon (and other) rows
#       may continue on following label-less rows (multiple weapons per ship).
#       One ship ('RAFT') is laid out VERTICALLY in a sidebar (label col H / value
#       col I) with no Cost header — handled specially (see AMBIGUITIES output).
#   - 'Power Plants' / 'Shields' / 'Quantum Drives' / 'Radars' / 'Coolers' :
#       [Model, Class, Size, Grade, Sell Price]. Manufacturer group-header rows
#       (model only, all other cells blank) are skipped.
#   - 'Weapons' : [Name, Size, Type, Sell Price, ...] (a side reference table in
#       cols E-I is ignored). No Class/Grade columns.
#   - 'Cheat Sheet' : haulers (cols A 'Hauling Ship' / B 'Grid Space') and the
#       material rates RMC/SCU (I16) and CMAT/SCU (I17).
# ============================================================================

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "[fetch:salvage-reference] openpyxl is required. Install it with:\n"
        "    pip install openpyxl"
    )

import os

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "electron" / "data" / "salvage-reference.json"

# The source worksheet path is supplied by the operator (it is never bundled and
# contains no personal data, but we keep no machine-specific path in the repo).
# Pass it as the first CLI argument, or set SALVAGE_WORKSHEET in the environment:
#     python scripts/fetch-salvage-reference.py "C:\\path\\to\\worksheet.xlsx"
#     SALVAGE_WORKSHEET="..." python scripts/fetch-salvage-reference.py

# Component sheets that share the [Model, Class, Size, Grade, Sell Price] layout,
# mapped to the component `type` the app uses.
STD_COMPONENT_SHEETS = {
    "Power Plants": "powerplant",
    "Shields": "shield",
    "Quantum Drives": "quantumdrive",
    "Radars": "radar",
    "Coolers": "cooler",
}

# Salvage Tracking attribute-row labels -> ship component keys.
TRACKING_COMPONENT_LABELS = {
    "Power Plant": "powerplant",
    "Shield": "shield",
    "Quantum Drive": "quantumdrive",
    "Cooler": "cooler",
    "Radar": "radar",
}
TRACKING_LABELS = set(TRACKING_COMPONENT_LABELS) | {
    "Cargo SCU",
    "CMAT",
    "Weapon",
    "Ship",
}

_COST_RE = re.compile(r"Cost\s*:\s*([\d,]+)\s*\(([\d,]+)\)")


def _parse_cost(text):
    """'Cost : 300 (150)' -> (300, 150). None if it doesn't match."""
    m = _COST_RE.match(str(text).strip())
    if not m:
        return None
    return int(m.group(1).replace(",", "")), int(m.group(2).replace(",", ""))


def _clean(v):
    """Normalize a cell: blank/whitespace -> None; strip strings."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def _num(v):
    """Return v if it is a finite number, else None."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    return None


def _str_or_none(v):
    c = _clean(v)
    return str(c) if c is not None else None


def _new_ship(name, cost_tier, claim_cost_org):
    return {
        "name": name,
        "costTier": cost_tier,
        "claimCost": None,
        "claimCostOrg": claim_cost_org,
        "cmat": None,
        "cargoScu": None,
        "components": {
            "powerplant": None,
            "shield": None,
            "quantumdrive": None,
            "cooler": None,
            "radar": None,
            "weapons": [],
        },
    }


def parse_salvage_tracking(ws, ambiguities):
    """Parse the 'Salvage Tracking' sheet -> list of ship dicts."""
    max_row, max_col = ws.max_row, ws.max_column

    # Locate the single vertical sidebar (a 'Ship' cell NOT in column A whose
    # right neighbor is a ship name). Exclude its cell box from the horizontal
    # parse so its labels are not mistaken for ship names.
    sidebar_cells = set()
    sidebar_anchor = None
    for r in range(1, max_row + 1):
        for c in range(2, max_col + 1):
            if ws.cell(r, c).value == "Ship":
                right = ws.cell(r, c + 1).value
                if isinstance(right, str) and right.strip():
                    sidebar_anchor = (r, c)
    if sidebar_anchor:
        sr0, sc0 = sidebar_anchor
        for dr in range(0, 12):
            sidebar_cells.add((sr0 + dr, sc0))
            sidebar_cells.add((sr0 + dr, sc0 + 1))

    # Cost-tier block headers in column A.
    headers = []
    for r in range(1, max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().startswith("Cost"):
            pc = _parse_cost(v)
            if pc:
                headers.append((r, pc[0], pc[1]))

    ships = []
    for bi, (hr, cost, org) in enumerate(headers):
        end = headers[bi + 1][0] - 1 if bi + 1 < len(headers) else max_row
        ship_row = next(
            (r for r in range(hr, end + 1) if ws.cell(r, 1).value == "Ship"), None
        )
        if ship_row is None:
            continue
        cols = {}
        for c in range(2, max_col + 1):
            if (ship_row, c) in sidebar_cells:
                continue
            nm = ws.cell(ship_row, c).value
            if isinstance(nm, str) and nm.strip():
                cols[c] = nm.strip()
        per_col = {c: _new_ship(cols[c], cost, org) for c in cols}

        cur = None
        for r in range(ship_row + 1, end + 1):
            lab = ws.cell(r, 1).value
            if isinstance(lab, str) and lab.strip().startswith("Cost"):
                break
            if isinstance(lab, str) and lab.strip() in TRACKING_LABELS:
                cur = lab.strip()
            elif isinstance(lab, str) and lab.strip():
                cur = None  # an unrecognized label ends continuation
            if cur is None or cur == "Ship":
                continue
            for c in cols:
                if (r, c) in sidebar_cells:
                    continue
                v = _clean(ws.cell(r, c).value)
                if v is None:
                    continue
                ship = per_col[c]
                if cur == "Cargo SCU":
                    ship["cargoScu"] = _num(v)
                elif cur == "CMAT":
                    ship["cmat"] = _num(v)
                elif cur == "Weapon":
                    ship["components"]["weapons"].append(str(v))
                else:
                    ship["components"][TRACKING_COMPONENT_LABELS[cur]] = str(v)
        for c in sorted(per_col):
            ships.append(per_col[c])

    # Vertical sidebar ship (RAFT): no Cost header of its own.
    if sidebar_anchor:
        sr0, sc0 = sidebar_anchor
        name = ws.cell(sr0, sc0 + 1).value.strip()
        ship = _new_ship(name, 500, 250)
        ambiguities.append(
            f"Ship '{name}' is laid out in a vertical sidebar with no 'Cost' "
            f"header; assigned costTier=500 / claimCostOrg=250 from its SC 4.8 "
            f"claim value and visual placement beside the 500 tier block."
        )
        for dr in range(0, 12):
            lab = ws.cell(sr0 + dr, sc0).value
            val = _clean(ws.cell(sr0 + dr, sc0 + 1).value)
            if not isinstance(lab, str) or val is None:
                continue
            lab = lab.strip()
            if lab == "Cargo SCU":
                ship["cargoScu"] = _num(val)
            elif lab == "CMAT":
                ship["cmat"] = _num(val)
            elif lab == "Weapon":
                ship["components"]["weapons"].append(str(val))
            elif lab in TRACKING_COMPONENT_LABELS:
                ship["components"][TRACKING_COMPONENT_LABELS[lab]] = str(val)
        ships.append(ship)

    return ships


def parse_components(wb):
    """Parse the six component sheets -> list of component dicts."""
    comps = []
    for sheet, ctype in STD_COMPONENT_SHEETS.items():
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            model = ws.cell(r, 1).value
            cls = ws.cell(r, 2).value
            size = ws.cell(r, 3).value
            grade = ws.cell(r, 4).value
            price = ws.cell(r, 5).value
            if model is None or not str(model).strip():
                continue
            # Manufacturer group-header row: model only, everything else blank.
            if cls is None and size is None and grade is None and price is None:
                continue
            comps.append(
                {
                    "type": ctype,
                    "model": str(model).strip(),
                    "class": _str_or_none(cls),
                    "size": _num(size),
                    "grade": _str_or_none(grade),
                    "sellPrice": _num(price),
                }
            )

    # Weapons sheet: [Name, Size, Type, Sell Price, ...]. No Class/Grade.
    ws = wb["Weapons"]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name is None or not str(name).strip():
            continue
        comps.append(
            {
                "type": "weapon",
                "model": str(name).strip(),
                "class": _str_or_none(ws.cell(r, 3).value),  # weapon Type
                "size": _num(ws.cell(r, 2).value),
                "grade": None,
                "sellPrice": _num(ws.cell(r, 4).value),
            }
        )
    return comps


def parse_cheat_sheet(ws):
    """Parse 'Cheat Sheet' -> (material_prices, haulers)."""
    haulers = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        grid = ws.cell(r, 2).value
        if isinstance(name, str) and name.strip() and isinstance(grid, (int, float)):
            haulers.append({"name": name.strip(), "gridScu": grid})

    rmc = _num(ws.cell(16, 9).value)  # I16
    cmat = _num(ws.cell(17, 9).value)  # I17
    material_prices = {
        "rmcPerScu": rmc if rmc is not None else 7200,
        "cmatPerScu": cmat if cmat is not None else 12000,
    }
    return material_prices, haulers


def main():
    worksheet_path = (
        sys.argv[1] if len(sys.argv) > 1 else os.environ.get("SALVAGE_WORKSHEET")
    )
    if not worksheet_path:
        sys.exit(
            "[fetch:salvage-reference] no worksheet given. Pass a path as the "
            "first argument or set SALVAGE_WORKSHEET in the environment."
        )
    if not Path(worksheet_path).exists():
        sys.exit(f"[fetch:salvage-reference] worksheet not found: {worksheet_path}")

    print(f"[fetch:salvage-reference] reading {worksheet_path}")
    wb = openpyxl.load_workbook(worksheet_path, data_only=True)

    ambiguities = []
    ships = parse_salvage_tracking(wb["Salvage Tracking"], ambiguities)
    components = parse_components(wb)
    material_prices, haulers = parse_cheat_sheet(wb["Cheat Sheet"])

    if not ships or not components or not haulers:
        sys.exit(
            f"[fetch:salvage-reference] refusing to write a thin snapshot "
            f"(ships={len(ships)}, components={len(components)}, "
            f"haulers={len(haulers)})."
        )

    snapshot = {
        "fetchedAt": 0,  # deterministic — this is static reference data
        "source": "Community SC 4.8 salvage worksheet (sanitized; game data only)",
        "materialPrices": material_prices,
        "ships": ships,
        "components": components,
        "haulers": haulers,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(snapshot, indent=2) + "\n", encoding="utf-8")

    by_type = {}
    for c in components:
        by_type[c["type"]] = by_type.get(c["type"], 0) + 1

    print(f"[fetch:salvage-reference] wrote {OUT_PATH}")
    print(f"  ships:      {len(ships)}")
    print(f"  components: {len(components)}  {by_type}")
    print(f"  haulers:    {len(haulers)}")
    print(f"  materials:  RMC={material_prices['rmcPerScu']}/SCU "
          f"CMAT={material_prices['cmatPerScu']}/SCU")
    if ambiguities:
        print("  AMBIGUITIES:")
        for a in ambiguities:
            print(f"    - {a}")


if __name__ == "__main__":
    main()
